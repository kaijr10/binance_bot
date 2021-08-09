# load selenium components
import time
import os
import pandas as pd
import argparse
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


class WebTable(object):

    def __init__(self, table, wait):
        self.table = table
        self.wait = wait
    
    def get_count_row(self):
        """
        Get the row count
        """
        return len(self.table.find_elements_by_tag_name("tr")) - 1;
    
    def row_data(self, row_number):
        """
        Get row data
        """
        row_data = []
        if row_number == 0:
            raise Exception("Row number must be greater than 0")
        for num in range(2, row_number + 1):
            try:
                self.wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div[2]/div[7]/div/div[1]/div/div/div/table/tbody/tr[" + str(num) + "]")))
                # sleep to wait for data is rendered
                time.sleep(1)
                row = self.table.find_elements_by_xpath("/html/body/div/div/div[2]/div[2]/div[7]/div/div[1]/div/div/div/table/tbody/tr[" + str(num) + "]")
                # Get each item of each symbol
                rows = row[0].text.split("\n")
                symbol_data = {}
                symbol_data['symbol'] = rows[0]
                symbol_data['size'] = rows[1]
                symbol_data['entry_price'] = rows[2]
                # append to row data
                row_data.append(symbol_data)
            except Exception as e:
                pass

        return row_data

class ExcelSaver(object):

    def __init__(self, data, trader_name):
        self.data = data
        self.file_name = trader_name + ".xlsx"
    
    def write_excel(self, truncate_sheet=False, sheet_name='Sheet1', startrow=None):
        if self.data:
            # save to output file
            df = pd.DataFrame(self.data)
            if not os.path.isfile(self.file_name):
                df.to_excel(self.file_name, index=None, sheet_name='Sheet1', startrow=0)
            else:
                # append it
                writer = pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a')

                # try to open an existing workbook
                writer.book = load_workbook(self.file_name)
                
                # get the last row in the existing Excel sheet
                # if it was not specified explicitly
                if startrow is None and sheet_name in writer.book.sheetnames:
                    startrow = writer.book[sheet_name].max_row

                # truncate sheet
                if truncate_sheet and sheet_name in writer.book.sheetnames:
                    # index of [sheet_name] sheet
                    idx = writer.book.sheetnames.index(sheet_name)
                    # remove [sheet_name]
                    writer.book.remove(writer.book.worksheets[idx])
                    # create an empty sheet [sheet_name] using old index
                    writer.book.create_sheet(sheet_name, idx)
                
                # copy existing sheets
                writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

                if startrow is None:
                    startrow = 0

                # write out the new sheet
                df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=None)

                # save the workbook
                writer.save()
    
    def get_data(self):
        return self.data

class Driver(object):

    def __init__(self, url):
        self.url = url
        #url = "https://www.binance.com/en/futures-activity/leaderboard?type=myProfile&encryptedUid=D64DDD2177FA081E3F361F70C703A562"
        self.driver = webdriver.Firefox()
        self.driver.get(url)
    
    def get_driver(self):
        return self.driver

    def close(self):
        self.driver.close()
    
    def refresh(self):
        self.driver.refresh()


class Strategy(object):

    def __init__(self, args):
        self.url = args.url
        self.driver = Driver(self.url).get_driver()
    
    def main(self):
        old_data = []
        while True:
            # waiting for the browser redered successful
            wait = WebDriverWait(self.driver, 20)

            # Get trader name as output file
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".css-1kmpww2")))
            trader_name = self.driver.find_elements_by_css_selector(".css-1kmpww2")
            trader_name = trader_name[0].text
            # find tab position click it 
            element = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"tab-MYPOSITIONS\"]")))
            element.click()

            # Find the table to extract the data
            wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div[2]/div[7]/div/div[1]/div/div/div/table")))
            w = WebTable(self.driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div[7]/div/div[1]/div/div/div/table"), wait)

            # Get rows data
            row_data = w.row_data(w.get_count_row())
            
            print(row_data)
            if self.compare_data(old_data, row_data):
                # Save Data to excel
                excel_saver = ExcelSaver(self.compare_data(old_data, row_data), trader_name)
                excel_saver.write_excel()
                old_data = row_data

            # sleep 30s
            print("Interval 30s to reload the page")
            time.sleep(5)

            # refresh the page
            self.driver.refresh()
    
    @staticmethod
    def compare_data(old_data, new_data):
        result_data = []
        if new_data:
            if old_data:
                for ele_old in old_data:
                    for ele_new in new_data:
                        if ele_old['symbol'] == ele_new['symbol']:
                            if ele_old['size'] != ele_new['size'] or ele_old['entry_price'] != ele_new['entry_price']:
                                result.append(ele_new)
            else:
                return new_data
        
        print(result_data)
        return result_data

def parser_args():
    # initiate parser
    parser = argparse.ArgumentParser(description="The binance bot to get price and entry price of symbols")

    # add arguments
    parser.add_argument('-i', '--url', type=str, required=True, help="The url of trader")
    
    args = parser.parse_args()
    
    return args

if __name__ == '__main__':
    strategy = Strategy(parser_args())
    strategy.main()
